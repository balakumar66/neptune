"""
Data I/O module for reading URLs and writing results.
"""

import pandas as pd
from typing import List, Optional
from pathlib import Path
import logging

from scraper import Section

logger = logging.getLogger(__name__)


def read_urls_from_csv(filepath: str, url_column: str = 'url') -> List[str]:
    """
    Read URLs from a CSV file.
    
    Args:
        filepath: Path to the CSV file
        url_column: Name of the column containing URLs
        
    Returns:
        List of URL strings
    """
    urls, _ = read_urls_with_metadata(filepath, url_column)
    return urls


def read_urls_with_metadata(filepath: str, url_column: str = 'url') -> tuple:
    """
    Read URLs from a CSV file along with any additional metadata columns.
    
    Args:
        filepath: Path to the CSV file
        url_column: Name of the column containing URLs
        
    Returns:
        Tuple of (List of URL strings, Dict mapping URL to metadata dict)
    """
    try:
        df = pd.read_csv(filepath)
        
        # Try to find the URL column (case-insensitive)
        columns_lower = {col.lower(): col for col in df.columns}
        
        if url_column.lower() in columns_lower:
            actual_column = columns_lower[url_column.lower()]
        elif 'url' in columns_lower:
            actual_column = columns_lower['url']
        elif 'link' in columns_lower:
            actual_column = columns_lower['link']
        elif 'website' in columns_lower:
            actual_column = columns_lower['website']
        else:
            # Use first column as fallback
            actual_column = df.columns[0]
            logger.warning(f"URL column not found, using first column: {actual_column}")
        
        # Build metadata dictionary (all columns except the URL column)
        url_metadata = {}
        other_columns = [col for col in df.columns if col != actual_column]
        
        for _, row in df.iterrows():
            url = str(row[actual_column]).strip()
            if url:
                metadata = {col: row[col] for col in other_columns}
                url_metadata[url] = metadata
        
        urls = df[actual_column].dropna().astype(str).tolist()
        
        # Clean URLs
        urls = [url.strip() for url in urls if url.strip()]
        
        logger.info(f"Read {len(urls)} URLs from {filepath}")
        return urls, url_metadata
        
    except Exception as e:
        logger.error(f"Failed to read CSV: {e}")
        raise


def sections_to_dataframe(sections: List[Section], url_metadata: dict = None) -> pd.DataFrame:
    """
    Convert a list of Section objects to a DataFrame.
    
    Args:
        sections: List of Section objects
        url_metadata: Optional dict mapping URL to metadata dict from input CSV
        
    Returns:
        DataFrame with section data
    """
    data = []
    for section in sections:
        row = {
            'url': section.url,
            'page_title': getattr(section, 'page_title', '') or '',
            'meta_description': getattr(section, 'meta_description', '') or '',
            'canonical_url': getattr(section, 'canonical_url', '') or '',
            'section_title': section.section_title,
            'section_level': section.section_level,
            'content': section.content,
            'ai_category': section.category or ''
        }
        
        # Add metadata from original CSV if available
        if url_metadata and section.url in url_metadata:
            for key, value in url_metadata[section.url].items():
                # Prefix with 'source_' to distinguish from extracted data
                row[f'source_{key.lower().replace(" ", "_")}'] = value
        
        data.append(row)
    
    df = pd.DataFrame(data)
    
    # Reorder columns: url, source columns, meta columns, section data
    if url_metadata:
        source_cols = [col for col in df.columns if col.startswith('source_')]
        meta_cols = ['page_title', 'meta_description', 'canonical_url']
        section_cols = ['section_title', 'section_level', 'content', 'ai_category']
        df = df[['url'] + source_cols + meta_cols + section_cols]
    else:
        meta_cols = ['page_title', 'meta_description', 'canonical_url']
        section_cols = ['section_title', 'section_level', 'content', 'ai_category']
        df = df[['url'] + meta_cols + section_cols]
    
    return df


def write_output(df: pd.DataFrame, filepath: str, add_separator: bool = True) -> None:
    """
    Write DataFrame to CSV or Excel based on file extension.
    
    Args:
        df: DataFrame to write
        filepath: Output file path (.csv or .xlsx)
        add_separator: If True, add row index and visual separator between records
    """
    path = Path(filepath)
    
    # Create a copy to avoid modifying original
    output_df = df.copy()
    
    if add_separator:
        # Add row number as first column for easy reference
        output_df.insert(0, 'row_no', range(1, len(output_df) + 1))
        
        # Add separator line between content from different URLs
        # Clean up content field - replace internal newlines with a marker
        if 'content' in output_df.columns:
            output_df['content'] = output_df['content'].apply(
                lambda x: str(x).replace('\n', ' | ').strip() if pd.notna(x) else ''
            )
    
    if path.suffix.lower() == '.xlsx':
        # For Excel, add formatting
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            output_df.to_excel(writer, index=False, sheet_name='Sections')
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Sections']
            
            # Auto-adjust column widths (approximate)
            for idx, col in enumerate(output_df.columns):
                max_length = max(
                    output_df[col].astype(str).map(len).max(),
                    len(col)
                )
                # Cap at 50 for content columns, 30 for others
                if col in ['content', 'meta_description']:
                    max_length = min(max_length, 80)
                else:
                    max_length = min(max_length, 40)
                worksheet.column_dimensions[chr(65 + idx) if idx < 26 else 'A'].width = max_length + 2
            
            # Add alternating row colors and borders for readability
            from openpyxl.styles import PatternFill, Border, Side
            
            light_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
            thin_border = Border(
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            # Track URL changes for visual grouping
            prev_url = None
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(output_df) + 1), start=2):
                # Get URL from the row (column index depends on structure)
                url_col_idx = list(output_df.columns).index('url') if 'url' in output_df.columns else 1
                current_url = output_df.iloc[row_idx - 2]['url'] if row_idx - 2 < len(output_df) else None
                
                # Add thick border when URL changes (new page)
                if prev_url and current_url != prev_url:
                    thick_border = Border(
                        top=Side(style='medium', color='4472C4')
                    )
                    for cell in row:
                        cell.border = thick_border
                
                prev_url = current_url
                
    else:
        # For CSV, add separator markers
        output_df.to_csv(filepath, index=False)
    
    logger.info(f"Wrote {len(output_df)} rows to {filepath}")


def create_summary_report(df: pd.DataFrame) -> pd.DataFrame:
    """
    Create a summary report of categories by URL.
    
    Args:
        df: DataFrame with extracted sections
        
    Returns:
        Summary DataFrame
    """
    if 'category' not in df.columns or df['category'].isna().all():
        return pd.DataFrame()
    
    summary = df.groupby(['url', 'category']).agg({
        'section_title': 'count',
        'content': lambda x: sum(len(str(c)) for c in x)
    }).reset_index()
    
    summary.columns = ['url', 'category', 'section_count', 'total_content_length']
    
    return summary
